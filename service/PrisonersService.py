import tempfile


import os
import openpyxl
from datetime import datetime, date
from fastapi import UploadFile
from dao.repository.PrisonersRepository import *

CHUNK_SIZE = 1024 * 1024  # 1 MB chunks

class PrisonersService:
    def __init__(self, db: Session):
        self.db = db
        self.buffer = None
        self.filename = None
        self.expected_size = None
        self.received_size = 0
        self.prisonersRepo = PrisonersRepository(db)

    @staticmethod
    def is_header_row(row):
        """Detect header rows by common Azerbaijani column names."""
        text = " ".join([str(c).lower() for c in row if c])
        keywords = ["adı", "soyadı", "ata", "fin", "doğum", "ş.v", "modul", "müəssisə"]

        return any(k in text for k in keywords)

    async def save_real_excel_stream(self, file: UploadFile):
        # Create temp file on disk
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        tmp_name = tmp.name

        # Write chunks
        while chunk := await file.read(CHUNK_SIZE):
            tmp.write(chunk)

        tmp.close()

        prisoners_batch = []

        try:
            workbook = openpyxl.load_workbook(tmp_name, read_only=True, data_only=True)

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                for row in sheet.iter_rows(values_only=True):

                    # 1. Skip empty rows
                    if not row or not row[0]:
                        continue

                    # 2. Skip header rows
                    if self.is_header_row(row):
                        continue

                    # 3. Extract name parts
                    adi = row[1] if len(row) > 0 else None
                    soyadi = row[2] if len(row) > 1 else None
                    ata_adi = row[3] if len(row) > 2 else None

                    full_name = " ".join([str(x) for x in [adi, soyadi, ata_adi] if x])

                    # 4. Helper conversion functions
                    def to_int(value):
                        try:
                            return int(value)
                        except:
                            return None

                    def to_date(value):
                        if isinstance(value, (datetime, date)):
                            return value

                        if isinstance(value, str):
                            try:
                                # Parse DD-MM-YYYY
                                return datetime.strptime(value, "%d-%m-%Y").date()
                            except ValueError:
                                return None  # invalid date format

                        return None

                    # 5. Mapping to Prisoner fields
                    prisoner = {
                        "full_name": full_name,
                        "pin": row[4],
                        "serial_type_id": row[5],
                        "serial_number": row[6],
                        "birth_date": to_date(row[7]),
                        "organization_id": to_int(row[8]),
                        "module_id": to_int(row[9]),
                        "sub_module_id": to_int(row[10]),
                        "division": to_int(row[11]),
                        "personal_number": to_int(row[12]),
                        "article": row[13],
                        "article": row[14],
                        "start_date": to_date(row[15]),
                        "end_date": to_date(row[16]),
                        "short_term_permit": to_int(row[17]),
                        "long_term_permit": to_int(row[18]),
                        "visit_item": to_int(row[19]),
                        "parcel": to_int(row[20]),
                        "deleted": False,
                    }

                    prisoners_batch.append(prisoner)

                    # 6. Save in chunks of 500
                    if len(prisoners_batch) >= 500:
                        self.prisonersRepo.save_all_prisoners(prisoners_batch)
                        prisoners_batch = []

            # Save remaining rows
            if prisoners_batch:
                self.prisonersRepo.save_all_prisoners(prisoners_batch)

            workbook.close()

        finally:
            try:
                os.remove(tmp_name)
            except Exception as e:
                print("Cleanup error:", e)

        return {"status": "ok"}










