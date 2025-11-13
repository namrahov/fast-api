from starlette.exceptions import HTTPException
import io
from fastapi import UploadFile
from dao.repository.UserRepository import *
import openpyxl
import tempfile
import os

CHUNK_SIZE = 1024 * 1024  # 1 MB chunks


class UserService:
    def __init__(self, db: Session):
        self.db = db
        self.buffer = None
        self.filename = None
        self.expected_size = None
        self.received_size = 0
        self.userRepo = UserRepository(db)

    # -------------------------------
    # START UPLOAD
    # -------------------------------
    def start_upload(self, payload: dict):
        self.filename = payload["filename"]
        self.expected_size = payload["size"]
        self.buffer = io.BytesIO()
        self.received_size = 0

    # -------------------------------
    # WRITE CHUNK
    # -------------------------------
    def append_chunk(self, chunk: bytes):
        if self.buffer is None:
            raise Exception("Upload not started")

        self.buffer.write(chunk)
        self.received_size += len(chunk)
        return self.received_size

    # -------------------------------
    # FINISH UPLOAD
    # -------------------------------
    def finish_upload(self):
        if self.received_size != self.expected_size:
            raise Exception("Size mismatch")

        # Move to beginning of buffer
        self.buffer.seek(0)

        # Load Excel
        workbook = openpyxl.load_workbook(self.buffer)
        users = []

        for sheet in workbook.sheetnames:
            ws = workbook[sheet]

            for row in ws.iter_rows(values_only=True):
                if row and len(row) >= 2:
                    name, email = row[0], row[1]
                    if name and email:
                        users.append({"name": name, "email": email})

        # Save to DB
        self.userRepo.save_all_users(users)

        return len(users)

    async def upload_excel_file(self, file: UploadFile):
        if not file.filename.endswith((".xlsx", ".xls")):
            raise HTTPException(status_code=403, detail="Invalid file type. Only .xlsx or .xls allowed.")

        content = await file.read()
        workbook = openpyxl.load_workbook(io.BytesIO(content))

        users = []

        for sheet in workbook.sheetnames:
            ws = workbook[sheet]
        print(f"📄 Reading sheet: {sheet}")

        for row in ws.iter_rows(values_only=True):
            if row and len(row) >= 2:  # ensure at least two columns
                name, email = row[0], row[1]
                if name and email:  # skip empty rows
                    users.append({"name": name, "email": email})

        self.userRepo.save_all_users(users)

    async def save_excel_stream(self, file: UploadFile):
        # Create temp file on disk
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        tmp_name = tmp.name

        # Write chunks
        while chunk := await file.read(CHUNK_SIZE):
            tmp.write(chunk)

        tmp.close()  # Must close before openpyxl reads it on Windows

        users = []

        try:
            # IMPORTANT: open workbook like this
            workbook = openpyxl.load_workbook(tmp_name, read_only=True, data_only=True)

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                for row in sheet.iter_rows(values_only=True):
                    if row and len(row) >= 2:
                        name, email = row[0], row[1]
                        if name and email:
                            users.append({"name": name, "email": email})

                    if len(users) >= 500:
                        self.userRepo.save_all_users(users)
                        users = []

            if users:
                self.userRepo.save_all_users(users)

            # CLOSE WORKBOOK FIRST!
            workbook.close()

        finally:
            try:
                os.remove(tmp_name)  # NOW removal works
            except Exception as e:
                print("Cleanup error:", e)

        return {"status": "ok"}
