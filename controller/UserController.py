from main import app   # import the SAME app
from fastapi import FastAPI, UploadFile, File
from starlette import status
from starlette.exceptions import HTTPException
import io


import openpyxl
from service.UserService import save_all


@app.post("/upload-excel", status_code=status.HTTP_201_CREATED)
async def upload_excel(file: UploadFile = File(...)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=403, detail="Invalid file type. Only .xlsx or .xls allowed.")

    # Read uploaded file into memory
    content = await file.read()
    workbook = openpyxl.load_workbook(io.BytesIO(content))

    users = []  # collect all user rows

    for sheet in workbook.sheetnames:
        ws = workbook[sheet]
        print(f"📄 Reading sheet: {sheet}")

        for row in ws.iter_rows(values_only=True):
            if row and len(row) >= 2:  # ensure at least two columns
                name, email = row[0], row[1]
                if name and email:  # skip empty rows
                    users.append({"name": name, "email": email})

    if users:
        save_all(users)
        return {"message": f"✅ {len(users)} users saved successfully!"}
    else:
        return {"warning": "⚠️ No valid rows found to save."}

